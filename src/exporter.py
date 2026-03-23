"""
Exporter module: writes data to rotating XLSX files (500k rows each).
"""
import logging
import os
from typing import Optional, List, Any

logger = logging.getLogger(__name__)

MAX_ORIGINAL_PAIRS = int(os.environ.get("MAX_ORIGINAL_PAIRS", "6"))
MAX_ANALOG_PAIRS = int(os.environ.get("MAX_ANALOG_PAIRS", "12"))


def _pair_columns(prefix: str, pair_count: int) -> list:
    cols = []
    for idx in range(1, pair_count + 1):
        suffix = "" if idx == 1 else str(idx)
        cols.append(f"{prefix}_brand{suffix}")
        cols.append(f"{prefix}_number{suffix}")
    return cols


BASE_COLUMNS = [
    "source_section",
    "source_subsection",
    "source_url",
    "breadcrumb_path",
    "product_url",
    "product_id",
    "name",
    "brand",
    "part_number_display",
    "part_number_normalized",
    "price_pln",
    "vat_included",
    "characteristics",
    "fitment_make",
    "fitment_model",
    "fitment_model_type",
    "fitment_modification",
    "fitment_raw_line",
]

COLUMNS = (
    BASE_COLUMNS
    + _pair_columns("original", MAX_ORIGINAL_PAIRS)
    + _pair_columns("analog", MAX_ANALOG_PAIRS)
)

SHEET_NAME = "Fitment"


class RotatingXlsxWriter:
    """
    Writes rows to XLSX files, rotating to a new file every `row_limit` rows.
    File names: {base_name}_0001.xlsx, {base_name}_0002.xlsx, ...
    """

    def __init__(self, output_dir: str, base_name: str = "2407_fitment_PLN",
                 row_limit: int = 500_000):
        self.output_dir = output_dir
        self.base_name = base_name
        self.row_limit = row_limit
        self.file_index = 1
        self.row_count = 0
        self.total_rows = 0
        self._workbook = None
        self._worksheet = None
        self._current_path = None
        os.makedirs(output_dir, exist_ok=True)
        self._open_new_file()

    def _current_file_path(self) -> str:
        return os.path.join(self.output_dir, f"{self.base_name}_{self.file_index:04d}.xlsx")

    def _open_new_file(self):
        """Close current file and open a new one."""
        self._close()
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        self._current_path = self._current_file_path()
        self._workbook = openpyxl.Workbook(write_only=False)
        self._worksheet = self._workbook.active
        self._worksheet.title = SHEET_NAME

        # Write header with styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = self._worksheet.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Set column widths
        col_widths = {
            "source_section": 20, "source_subsection": 25, "source_url": 50,
            "breadcrumb_path": 40, "product_url": 50, "product_id": 12,
            "name": 40, "brand": 20, "part_number_display": 20,
            "part_number_normalized": 20, "price_pln": 12, "vat_included": 12,
            "characteristics": 90,
            "fitment_make": 20,
            "fitment_model": 25,
            "fitment_model_type": 25,
            "fitment_modification": 35,
            "fitment_raw_line": 50,
        }
        for idx in range(1, MAX_ORIGINAL_PAIRS + 1):
            suffix = "" if idx == 1 else str(idx)
            col_widths[f"original_brand{suffix}"] = 20
            col_widths[f"original_number{suffix}"] = 24
        for idx in range(1, MAX_ANALOG_PAIRS + 1):
            suffix = "" if idx == 1 else str(idx)
            col_widths[f"analog_brand{suffix}"] = 20
            col_widths[f"analog_number{suffix}"] = 24
        for col_idx, col_name in enumerate(COLUMNS, 1):
            self._worksheet.column_dimensions[get_column_letter(col_idx)].width = \
                col_widths.get(col_name, 20)

        # Freeze header row
        self._worksheet.freeze_panes = "A2"

        self.row_count = 0
        logger.info(f"Opened new output file: {self._current_path}")

    def _close(self):
        """Save and close the current workbook."""
        if self._workbook is not None:
            try:
                self._workbook.save(self._current_path)
                logger.info(f"Saved {self._current_path} ({self.row_count} data rows)")
            except Exception as e:
                logger.error(f"Error saving {self._current_path}: {e}")
            self._workbook = None
            self._worksheet = None

    def write_row(self, row_data: dict):
        """Write a single row. Rotates file if limit is reached."""
        if self.row_count >= self.row_limit:
            self.file_index += 1
            self._open_new_file()

        values = [row_data.get(col) for col in COLUMNS]
        self._worksheet.append(values)
        self.row_count += 1
        self.total_rows += 1

    def write_product(self, product_data, source_ctx: dict):
        """Write all fitment rows for a product."""
        base_row = {
            "source_section": source_ctx.get("source_section"),
            "source_subsection": source_ctx.get("source_subsection"),
            "source_url": source_ctx.get("source_url"),
            "breadcrumb_path": product_data.breadcrumb_path,
            "product_url": product_data.product_url,
            "product_id": product_data.product_id,
            "name": product_data.name,
            "brand": product_data.brand,
            "part_number_display": product_data.part_number_display,
            "part_number_normalized": product_data.part_number_normalized,
            "price_pln": product_data.price_pln,
            "vat_included": product_data.vat_included,
            "characteristics": getattr(product_data, "characteristics", None),
        }

        for fitment in product_data.fitment_rows:
            row = dict(base_row)
            row["fitment_make"] = fitment.make
            row["fitment_model"] = fitment.model
            row["fitment_model_type"] = getattr(fitment, "model_type", None)
            row["fitment_modification"] = getattr(fitment, "modification", None)
            row["fitment_raw_line"] = fitment.raw_line
            self.write_row(row)

    def finalize(self):
        """Close and finalize output."""
        self._close()
        logger.info(f"Export complete. Total rows: {self.total_rows}, files: {self.file_index}")

    def __del__(self):
        self._close()


class CsvWriter:
    """Optional CSV writer with same schema."""

    def __init__(self, output_dir: str, base_name: str = "2407_fitment_PLN",
                 row_limit: int = 500_000):
        import csv
        self.output_dir = output_dir
        self.base_name = base_name
        self.row_limit = row_limit
        self.file_index = 1
        self.row_count = 0
        self.total_rows = 0
        self._file = None
        self._writer = None
        os.makedirs(output_dir, exist_ok=True)
        self._open_new_file()

    def _open_new_file(self):
        if self._file:
            self._file.close()
        path = os.path.join(self.output_dir, f"{self.base_name}_{self.file_index:04d}.csv")
        self._file = open(path, "w", newline="", encoding="utf-8-sig")
        import csv
        self._writer = csv.DictWriter(self._file, fieldnames=COLUMNS, extrasaction="ignore")
        self._writer.writeheader()
        self.row_count = 0
        logger.info(f"Opened CSV: {path}")

    def write_row(self, row_data: dict):
        if self.row_count >= self.row_limit:
            self.file_index += 1
            self._open_new_file()
        self._writer.writerow({col: row_data.get(col, "") for col in COLUMNS})
        self.row_count += 1
        self.total_rows += 1

    def finalize(self):
        if self._file:
            self._file.close()
        logger.info(f"CSV export done. Total: {self.total_rows}")
